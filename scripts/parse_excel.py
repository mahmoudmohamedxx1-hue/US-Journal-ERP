"""
Parse the US Journal.xlsx workbook into structured JSON files
for the Pro ERP System.

Output: src/data/*.json
"""
import openpyxl
import json
import os
import re
from datetime import datetime, date

SRC = '/home/z/my-project/upload/US Journal.xlsx'
OUT = '/home/z/my-project/src/data'
os.makedirs(OUT, exist_ok=True)

# Arabic → English translations for account names (best-effort)
AR_EN = {
    'الاصول': 'Assets',
    'الاصول الثابتة': 'Fixed Assets',
    'أصول غير ملموسة': 'Intangible Assets',
    'مشروعات تحت التنفيذ': 'Projects Under Construction',
    'أصول أخري طويلة الأجل': 'Other Long-Term Assets',
    'المخزون': 'Inventory',
    'عملاء و أوراق قبض': 'Customers & Receivables',
    'مدينون و أرصدة مدينة أخرى': 'Debit & Other Debit Balances',
    'مستحق من أطراف ذات علاقة': 'Due from Related Parties',
    'أصول ضريبة جارية': 'Current Tax Assets',
    'النقدية بالخزينة ولدي البنوك': 'Cash & Cash Equivalents',
    'الخصوم وحقوق الملكية': 'Liabilities & Equity',
    'الخصوم طويلة الاجل': 'Long-Term Liabilities',
    'قروض طويلة الاجل': 'Long-Term Loans',
    'مخصصات': 'Provisions',
    'ايلاء ضريبية مؤجلة': 'Deferred Tax Liabilities',
    'الخصوم المتداولة': 'Current Liabilities',
    'دائنون و أوراق دفع': 'Suppliers & Payables',
    'مستحق لأطراف ذات علاقة': 'Due to Related Parties',
    'مستحقات أخرى': 'Other Payables',
    'ضرائب مستحقة': 'Taxes Payable',
    'حقوق الملكية': 'Equity',
    'رأس المال': 'Capital',
    'الأرباح المحتجزة': 'Retained Earnings',
    'الإيرادات': 'Revenue',
    'المصروفات': 'Expenses',
    'تكلفة الايراد': 'Cost of Revenue',
    'مصاريف إدارية و عمومية': 'G&A Expenses',
    'مصاريف بيع وتسويق': 'Sales & Marketing Expenses',
    'فرق عملة': 'Foreign Exchange',
    'مصاريف أخرى': 'Other Expenses',
    'مصاريف تمويلية': 'Finance Expenses',
    'فوائد دائنة': 'Credit Interest',
    'أرباح رأس مالية': 'Capital Gains',
    'ضريبة الدخل': 'Income Tax',
    'نصيب في الأرباح': 'Share in Profit',
    'قيد اثبات رأس المال': 'Capital introduction entry',
    'رصيد اول المدة': 'Opening balance',
    'اجهزة كمبيوتر وطابعات': 'Computers & Printers',
    'عهدة ا- عمر داوود': "Custody - Omar Dawood",
    'مصاريف مدفوعه مقدما': 'Prepaid Expenses',
}

MONTH_SHEETS = [
    ('jan 2026.....', 1, 'January'),
    ('Feb 2026.', 2, 'February'),
    ('Mar 2026', 3, 'March'),
    ('Apr 2026', 4, 'April'),
    ('May 2026', 5, 'May'),
    ('Jun 2026.', 6, 'June'),
    ('July 202.....', 7, 'July'),
    ('Aug 202....', 8, 'August'),
    ('Sep 202....', 9, 'September'),
    ('Oct 202....', 10, 'October'),
    ('Nov 202....', 11, 'November'),
    ('Des 202.....', 12, 'December'),
]


def tr(ar_text):
    """Translate Arabic to English if we have a mapping."""
    if ar_text is None:
        return None
    ar_text = str(ar_text).strip()
    if not ar_text:
        return None
    # Try direct mapping
    if ar_text in AR_EN:
        return AR_EN[ar_text]
    return ar_text


def to_num(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(',', ''))
    except (ValueError, TypeError):
        return 0


def to_iso_date(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return str(v)


def parse_chart_of_accounts(ws):
    """Parse the Chart of Account sheet into a hierarchical structure."""
    accounts = []
    current_main = None
    current_sub = None
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        code, name = row[0], row[1]
        if code is None and name is None:
            continue
        if isinstance(code, str) and not code.strip():
            continue
        try:
            code_int = int(code) if code is not None else None
        except (ValueError, TypeError):
            continue
        if code_int is None:
            continue

        name_ar = str(name).strip() if name else ''
        # Determine hierarchy level by code length
        if code_int < 100:
            # Main category (1, 2, 3...)
            account = {
                'code': str(code_int),
                'nameAr': name_ar,
                'nameEn': tr(name_ar),
                'level': 1,
                'parent': None,
                'type': 'main',
            }
            accounts.append(account)
            current_main = account
        elif code_int < 1000:
            # Sub-category (11, 12, 21, 22...)
            account = {
                'code': str(code_int),
                'nameAr': name_ar,
                'nameEn': tr(name_ar),
                'level': 2,
                'parent': current_main['code'] if current_main else None,
                'type': 'category',
            }
            accounts.append(account)
            current_sub = account
        else:
            # Detailed account (11001, 11002...)
            account = {
                'code': str(code_int),
                'nameAr': name_ar,
                'nameEn': tr(name_ar),
                'level': 3,
                'parent': current_sub['code'] if current_sub else None,
                'type': 'detail',
            }
            accounts.append(account)
    return accounts


def parse_monthly_journal(ws, month_num, month_name):
    """Parse a monthly journal sheet into entries."""
    # Read header row 1 to find account columns (col K onwards = index 10+)
    header_row = None
    account_cols = {}  # account_code -> col_index

    # First read all rows into list
    all_rows = []
    for r in ws.iter_rows(values_only=True):
        all_rows.append(r)

    if not all_rows:
        return []

    # Header row is row 1 (index 0)
    header = all_rows[0]
    for col_idx, cell in enumerate(header):
        if cell is None:
            continue
        try:
            code = int(cell)
            if code > 100:
                account_cols[code] = col_idx
        except (ValueError, TypeError):
            continue

    # Row 3 has the debit/credit labels: 'مدين' (debit) = col index 7, 'دائن' (credit) = col 8
    # We use these totals to detect entries with content
    DEBIT_COL = 7
    CREDIT_COL = 8
    ENTRY_COL = 0
    DATE_COL = 1
    DESC_COL = 2
    REF_IN_COL = 3
    REF_OUT_COL = 4
    INV_COL = 5
    CHK_COL = 6

    entries = []
    for row in all_rows[3:]:
        if not row or len(row) < 9:
            continue
        entry_no = row[ENTRY_COL]
        date_val = row[DATE_COL]
        desc = row[DESC_COL]
        debit = to_num(row[DEBIT_COL]) if DEBIT_COL < len(row) else 0
        credit = to_num(row[CREDIT_COL]) if CREDIT_COL < len(row) else 0

        # Skip rows without entry number AND without amounts
        if entry_no is None and debit == 0 and credit == 0:
            continue
        # Skip if all key fields empty
        if entry_no is None and not desc:
            continue

        # Extract line items per account
        line_items = []
        for acc_code, col_idx in account_cols.items():
            if col_idx >= len(row):
                continue
            val = to_num(row[col_idx])
            if val != 0:
                line_items.append({
                    'accountCode': str(acc_code),
                    'amount': val,
                })

        entries.append({
            'entryNo': entry_no if entry_no is not None else len(entries) + 1,
            'date': to_iso_date(date_val),
            'month': month_num,
            'monthName': month_name,
            'description': str(desc) if desc else '',
            'referenceIn': str(row[REF_IN_COL]) if REF_IN_COL < len(row) and row[REF_IN_COL] else None,
            'referenceOut': str(row[REF_OUT_COL]) if REF_OUT_COL < len(row) and row[REF_OUT_COL] else None,
            'invoiceNo': str(row[INV_COL]) if INV_COL < len(row) and row[INV_COL] else None,
            'checkNo': str(row[CHK_COL]) if CHK_COL < len(row) and row[CHK_COL] else None,
            'debit': debit,
            'credit': credit,
            'lineItems': line_items,
        })
    return entries


def parse_trial_balance(ws):
    """Parse the Trail Balance sheet."""
    rows = []
    for r in ws.iter_rows(min_row=1, max_col=10, values_only=True):
        code = r[0]
        name = r[1]
        op_debit = to_num(r[2])
        op_credit = to_num(r[3])
        mv_debit = to_num(r[4])
        mv_credit = to_num(r[5])
        tot_debit = to_num(r[6])
        tot_credit = to_num(r[7])
        bal_debit = to_num(r[8])
        bal_credit = to_num(r[9])

        if code is None and name is None:
            continue
        try:
            code_int = int(code) if code is not None else None
        except (ValueError, TypeError):
            code_int = None
        if code_int is None:
            continue

        rows.append({
            'code': str(code_int),
            'nameAr': str(name).strip() if name else '',
            'nameEn': tr(name),
            'openingDebit': op_debit,
            'openingCredit': op_credit,
            'movementDebit': mv_debit,
            'movementCredit': mv_credit,
            'totalDebit': tot_debit,
            'totalCredit': tot_credit,
            'balanceDebit': bal_debit,
            'balanceCredit': bal_credit,
        })
    return rows


def parse_balance_sheet(ws):
    """Parse the Balance Sheet sheet."""
    items = []
    for i, r in enumerate(ws.iter_rows(min_row=10, max_col=8, values_only=True)):
        label = r[0]
        note = r[1]
        cur_year = to_num(r[2])
        prev_year = to_num(r[4])

        if label is None or (isinstance(label, str) and not label.strip()):
            continue
        items.append({
            'label': str(label).strip(),
            'labelEn': tr(str(label).strip()) if label else '',
            'note': str(note) if note else None,
            'currentYear': cur_year,
            'previousYear': prev_year,
        })
    return items


def parse_income_statement(ws):
    """Parse the Income Statement sheet."""
    items = []
    for r in ws.iter_rows(min_row=1, max_col=8, values_only=True):
        label = r[0]
        cur_year = to_num(r[2]) if len(r) > 2 else 0
        prev_year = to_num(r[4]) if len(r) > 4 else 0
        note = r[1] if len(r) > 1 else None

        if label is None or (isinstance(label, str) and not label.strip()):
            continue
        items.append({
            'label': str(label).strip(),
            'labelEn': tr(str(label).strip()),
            'note': str(note) if note else None,
            'currentYear': cur_year,
            'previousYear': prev_year,
        })
    return items


def parse_simple_sheet(ws, max_rows=50, max_cols=10):
    """Parse a sheet as rows of values (for sub-ledgers)."""
    rows = []
    for r in ws.iter_rows(max_row=max_rows, max_col=max_cols, values_only=True):
        if all(c is None for c in r):
            continue
        rows.append([str(c) if c is not None else '' for c in r])
    return rows


def main():
    print(f'Loading workbook: {SRC}')
    wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)

    # 1. Chart of Accounts
    print('Parsing Chart of Account...')
    coa = parse_chart_of_accounts(wb['Chart of Account'])
    with open(f'{OUT}/accounts.json', 'w', encoding='utf-8') as f:
        json.dump(coa, f, ensure_ascii=False, indent=2)
    print(f'  {len(coa)} accounts')

    # 2. Monthly journals
    print('Parsing monthly journals...')
    all_entries = []
    for sheet_name, month_num, month_name in MONTH_SHEETS:
        try:
            ws = wb[sheet_name]
            entries = parse_monthly_journal(ws, month_num, month_name)
            for e in entries:
                all_entries.append(e)
            print(f'  {month_name}: {len(entries)} entries')
        except Exception as ex:
            print(f'  ERROR parsing {sheet_name}: {ex}')
    with open(f'{OUT}/journal-entries.json', 'w', encoding='utf-8') as f:
        json.dump(all_entries, f, ensure_ascii=False, indent=2)
    print(f'  Total: {len(all_entries)} entries')

    # 3. Trial Balance
    print('Parsing Trial Balance...')
    tb = parse_trial_balance(wb['Trail Balance'])
    with open(f'{OUT}/trial-balance.json', 'w', encoding='utf-8') as f:
        json.dump(tb, f, ensure_ascii=False, indent=2)
    print(f'  {len(tb)} rows')

    # 4. Balance Sheet
    print('Parsing Balance Sheet...')
    bs = parse_balance_sheet(wb['Balance Sheet'])
    with open(f'{OUT}/balance-sheet.json', 'w', encoding='utf-8') as f:
        json.dump(bs, f, ensure_ascii=False, indent=2)
    print(f'  {len(bs)} items')

    # 5. Income Statement
    print('Parsing Income Statement...')
    inc = parse_income_statement(wb['Incom Statment'])
    with open(f'{OUT}/income-statement.json', 'w', encoding='utf-8') as f:
        json.dump(inc, f, ensure_ascii=False, indent=2)
    print(f'  {len(inc)} items')

    # 6. Cash Flow
    print('Parsing Cash Flow...')
    try:
        cf = parse_income_statement(wb['Cash Flow Statment'])
        with open(f'{OUT}/cash-flow.json', 'w', encoding='utf-8') as f:
            json.dump(cf, f, ensure_ascii=False, indent=2)
        print(f'  {len(cf)} items')
    except Exception as ex:
        print(f'  ERROR: {ex}')

    # 7. Sub-ledgers (parse as simple tables)
    sub_ledgers = [
        'Fixed Assets', 'Intengable Assets', 'Inventory', 'Custamer ',
        'Suppliers', 'Cash & Cash eq', 'Capital', 'Retained Earnings',
        'Revanue', 'Cost', 'G&A', 'Sales& Marketing xp',
        'Finance Expenses', 'Provisions',
    ]
    print(f'Parsing sub-ledgers...')
    sub_ledger_data = {}
    for sheet_name in sub_ledgers:
        try:
            ws = wb[sheet_name]
            rows = parse_simple_sheet(ws, max_rows=40, max_cols=15)
            sub_ledger_data[sheet_name] = rows
            print(f'  {sheet_name}: {len(rows)} rows')
        except Exception as ex:
            print(f'  ERROR {sheet_name}: {ex}')
    with open(f'{OUT}/sub-ledgers.json', 'w', encoding='utf-8') as f:
        json.dump(sub_ledger_data, f, ensure_ascii=False, indent=2)

    # 8. Compute summary metrics
    print('Computing summary metrics...')
    summary = {
        'totalEntries': len(all_entries),
        'totalDebit': sum(e['debit'] for e in all_entries),
        'totalCredit': sum(e['credit'] for e in all_entries),
        'monthsCovered': len(set(e['month'] for e in all_entries)),
        'year': 2026,
        'currency': 'EGP',
        'companyName': 'US Holdings - شركة',
        'entriesByMonth': {},
    }
    for m in range(1, 13):
        month_entries = [e for e in all_entries if e['month'] == m]
        summary['entriesByMonth'][m] = {
            'count': len(month_entries),
            'debit': sum(e['debit'] for e in month_entries),
            'credit': sum(e['credit'] for e in month_entries),
        }
    with open(f'{OUT}/summary.json', 'w', encoding='utf-8') as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    print('\n=== DONE ===')
    print(f'Output: {OUT}/')


if __name__ == '__main__':
    main()
